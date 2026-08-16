import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import Application


ACCENT = "250746"
LIGHT = "EEE9F4"
STATUSES = [
    "Nueva",
    "Analizada",
    "Documentos listos",
    "Aplicada",
    "Seguimiento",
    "Entrevista RH",
    "Entrevista técnica",
    "Prueba",
    "Oferta",
    "Rechazada",
    "Retirada",
    "Archivada",
]


def _style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 28


def build_workbook(db: Session) -> BytesIO:
    applications = db.scalars(
        select(Application)
        .options(selectinload(Application.interviews))
        .order_by(Application.created_at.desc())
    ).all()

    workbook = Workbook()
    applications_sheet = workbook.active
    applications_sheet.title = "Candidaturas"
    headers = [
        "ID", "Fecha registro", "Empresa", "Puesto", "Enlace", "Fuente", "Ubicación",
        "Modalidad", "Contrato", "Idioma", "Salario", "Afinidad", "Estado",
        "Fecha aplicación", "Próximo seguimiento", "Contacto", "Email contacto",
        "Competencias coincidentes", "Competencias faltantes", "CV", "Carta",
        "Resultado", "Motivo de rechazo", "Notas",
    ]
    applications_sheet.append(headers)
    for item in applications:
        analysis = json.loads(item.analysis_json or "{}")
        applications_sheet.append(
            [
                item.id, item.created_at, item.company, item.role, item.url, item.source,
                item.location, item.work_mode, item.contract_type, item.language, item.salary,
                item.fit_score, item.status, item.applied_at, item.next_follow_up,
                item.contact_name, item.contact_email,
                ", ".join(analysis.get("matched_skills", [])),
                ", ".join(analysis.get("missing_skills", [])),
                item.cv_path, item.letter_path, item.outcome, item.rejection_reason, item.notes,
            ]
        )
    _style_header(applications_sheet)
    applications_sheet.freeze_panes = "A2"
    applications_sheet.auto_filter.ref = applications_sheet.dimensions
    applications_sheet.column_dimensions["C"].width = 24
    applications_sheet.column_dimensions["D"].width = 32
    applications_sheet.column_dimensions["E"].width = 45
    for column in ["R", "S", "W", "X"]:
        applications_sheet.column_dimensions[column].width = 38
    status_validation = DataValidation(
        type="list", formula1="=Catalogos!$A$2:$A$13", allow_blank=True
    )
    applications_sheet.add_data_validation(status_validation)
    status_validation.add(f"M2:M{max(2, applications_sheet.max_row + 1000)}")

    interviews_sheet = workbook.create_sheet("Entrevistas")
    interviews_sheet.append(
        ["ID", "Candidatura", "Empresa", "Puesto", "Tipo", "Fecha", "Entrevistador", "Reunión", "Resultado", "Notas"]
    )
    for application in applications:
        for interview in application.interviews:
            interviews_sheet.append(
                [
                    interview.id, application.id, application.company, application.role,
                    interview.kind, interview.scheduled_at, interview.interviewer,
                    interview.meeting_url, interview.result, interview.notes,
                ]
            )
    _style_header(interviews_sheet)
    interviews_sheet.freeze_panes = "A2"
    interviews_sheet.column_dimensions["C"].width = 24
    interviews_sheet.column_dimensions["D"].width = 30
    interviews_sheet.column_dimensions["H"].width = 40
    interviews_sheet.column_dimensions["J"].width = 40

    dashboard = workbook.create_sheet("Dashboard")
    dashboard.append(["Indicador", "Valor"])
    dashboard.append(["Total candidaturas", len(applications)])
    dashboard.append(["Aplicadas", sum(item.status not in {"Nueva", "Analizada", "Documentos listos"} for item in applications)])
    dashboard.append(["Con entrevista", sum(bool(item.interviews) for item in applications)])
    dashboard.append(["Ofertas", sum(item.status == "Oferta" for item in applications)])
    dashboard.append(["Rechazadas", sum(item.status == "Rechazada" for item in applications)])
    dashboard.append(["Afinidad promedio", round(sum(item.fit_score or 0 for item in applications) / max(1, len(applications)), 1)])
    _style_header(dashboard)
    dashboard.column_dimensions["A"].width = 28
    dashboard.column_dimensions["B"].width = 15
    chart = BarChart()
    chart.title = "Embudo de candidaturas"
    chart.y_axis.title = "Cantidad"
    chart.add_data(Reference(dashboard, min_col=2, min_row=2, max_row=6), titles_from_data=False)
    chart.set_categories(Reference(dashboard, min_col=1, min_row=2, max_row=6))
    chart.height = 7
    chart.width = 13
    dashboard.add_chart(chart, "D2")

    catalogs = workbook.create_sheet("Catalogos")
    catalogs.append(["Estados"])
    for status in STATUSES:
        catalogs.append([status])
    _style_header(catalogs)
    catalogs.sheet_state = "hidden"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

